from django.shortcuts import render, redirect
from .models import Docente, Curso, Servidor, Documento, DeclaracaoEmitida
from .forms import DocenteForm, CursoForm, ServidorForm, DeclaracaoEmitidaForm
from django.http import FileResponse
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from django.contrib.auth.decorators import login_required
from django.contrib.auth.decorators import permission_required
from django.shortcuts import render, get_object_or_404, redirect
import re

def format_data_nascimento(data):
    # Formata a data de nascimento no formato DD/MM/YYYY
    return data.strftime('%d/%m/%Y')

def format_cpf(cpf):
    # Remove caracteres não numéricos e formata o CPF no formato XXX.XXX.XXX-XX
    cpf = re.sub('[^0-9]', '', cpf)
    return f'{cpf[:3]}.{cpf[3:6]}.{cpf[6:9]}-{cpf[9:]}'

def format_telefone(telefone):
    # Remove caracteres não numéricos e formata o telefone no formato (XX) XXXXX-XXXX
    telefone = re.sub('[^0-9]', '', telefone)
    return f'({telefone[:2]}) {telefone[2:7]}-{telefone[7:]}'



@login_required
def cadastro_docente(request):
    if request.method == 'POST':
        form = DocenteForm(request.POST)
        if form.is_valid():
            cleaned_data = form.cleaned_data
            cleaned_data['data_nascimento'] = format_data_nascimento(cleaned_data['data_nascimento'])
            cleaned_data['cpf'] = format_cpf(cleaned_data['cpf'])
            cleaned_data['telefone'] = format_telefone(cleaned_data['telefone'])
            form.save()
            return redirect('cadastro_docente')
    else:
        form = DocenteForm()
    
    cursos = Curso.objects.all()  # Obtenha todos os cursos do banco de dados
    docentes = Docente.objects.all()  # Obtenha todos os docentes do banco de dados

    return render(request, 'cadastro_docente.html', {'form': form, 'cursos': cursos, 'docentes': docentes})




@login_required
def cadastro_curso(request):
    if request.method == 'POST':
        form = CursoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('cadastro_curso')
    else:
        form = CursoForm()

    cursos = Curso.objects.all()  # Obtenha todos os cursos do banco de dados
    docentes = Docente.objects.order_by('nome')  # Obtenha todos os docentes do banco de dados

    return render(request, 'cadastro_curso.html', {'form': form, 'cursos': cursos, 'docentes': docentes})







def busca(request):
    cursos = Curso.objects.all()
    docentes = Docente.objects.all()

    if request.method == 'POST':
        curso_id = request.POST.get('curso_id')
        docente_id = request.POST.get('docente_id')

        if curso_id:
            cursos = cursos.filter(id=curso_id)
        
        if docente_id:
            cursos = cursos.filter(docente__id=docente_id)

        if 'export_excel' in request.POST:
            filename = export_to_excel(cursos)
            response = FileResponse(open(filename, 'rb'), as_attachment=True, filename='cursos.xlsx')
            return response

    return render(request, 'busca.html', {'cursos': cursos, 'docentes': docentes})



def export_to_excel(request):
    cursos = Curso.objects.all()

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    headers = ['Nome', 'Docente', 'Turma', 'Ano', 'Componente', 'Perfil', 'Carga Horária', 'Período']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = sheet[f'{col_letter}1']
        cell.value = header
        cell.font = openpyxl.styles.Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for row_num, curso in enumerate(cursos, 2):
        sheet[f'A{row_num}'] = curso.nome
        if curso.docente:
            sheet[f'B{row_num}'] = curso.docente.nome
        else:
            sheet[f'B{row_num}'] = ""
        sheet[f'C{row_num}'] = curso.turma
        sheet[f'D{row_num}'] = curso.ano
        sheet[f'E{row_num}'] = curso.componente
        sheet[f'F{row_num}'] = curso.perfil
        sheet[f'G{row_num}'] = curso.carga_horaria
        sheet[f'H{row_num}'] = curso.periodo

    for col_num in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_num)
        sheet.column_dimensions[col_letter].width = 15

    # Determine o caminho completo onde deseja salvar o arquivo
    caminho_completo = '/var/www/ngtes/ngtes/documentos/cursos.xlsx'

    # Salve o arquivo no caminho completo especificado
    workbook.save(caminho_completo)

    # Abra o arquivo e crie uma resposta para download
    with open(caminho_completo, 'rb') as file:
        response = FileResponse(file, as_attachment=True, filename='cursos.xlsx')

    return response




from django.db.models import Q

@login_required
def index(request):
    query = request.POST.get('query', '')

    if query == '.':
        cursos = Curso.objects.all()
    else:
        cursos = Curso.objects.filter(Q(nome__icontains=query) | Q(docente__nome__icontains=query))

    return render(request, 'index.html', {'cursos': cursos})


def perfil_docente(request, docente_id):
    # Lógica para obter os detalhes do docente com base no 'docente_id'
    # ...
    docente = Docente.objects.get(id=docente_id)
    
    return render(request, 'perfil_docente.html', {'docente': docente})
# views.py (do aplicativo)
from django.shortcuts import render, redirect
from django.contrib.auth.forms import UserCreationForm

def register_view(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('login')
    else:
        form = UserCreationForm()
    
    return render(request, 'register.html', {'form': form})


def editar_curso(request, pk):
    curso = get_object_or_404(Curso, pk=pk)
    if request.method == 'POST':
        form = CursoForm(request.POST,instance=curso)
        if form.is_valid():
            form.save()
            return redirect('cadastro_curso')
    else:
        form = CursoForm(instance=curso)

    cursos = Curso.objects.all()  # Obtenha todos os cursos do banco de dados
    docentes = Docente.objects.all()  # Obtenha todos os docentes do banco de dados

    return render(request, 'cadastro_curso.html', {'form': form, 'cursos': cursos, 'docentes': docentes})

def editar_docente(request, pk):
    docentes = get_object_or_404(Docente, pk=pk)
    if request.method == 'POST':
        form = DocenteForm(request.POST,instance=docentes)
        if form.is_valid():
            form.save()
            return redirect('cadastro_docente')
    else:
        form = DocenteForm(instance=docentes)

    cursos = Curso.objects.all()  # Obtenha todos os cursos do banco de dados
    docentes = Docente.objects.all()  # Obtenha todos os docentes do banco de dados

    return render(request, 'cadastro_docente.html', {'form': form, 'cursos': cursos, 'docentes': docentes})


from django.db.models import Q

def cadastro_servidor(request):
    sort_field = request.GET.get('sort')  # Obtém o parâmetro "sort" da URL

    if request.method == 'POST':
        form = ServidorForm(request.POST)
        if form.is_valid():
            cleaned_data = form.cleaned_data
            cleaned_data['data_nascimento'] = format_data_nascimento(cleaned_data['data_nascimento'])
            cleaned_data['cpf'] = format_cpf(cleaned_data['cpf'])
            cleaned_data['telefone'] = format_telefone(cleaned_data['telefone'])
            form.save()
            return redirect('cadastro_servidor')
    else:
        form = ServidorForm()

    # Obtenha todos os servidores
    servidores = Servidor.objects.all()

    # Ordenação
    if sort_field:
        servidores = servidores.order_by(sort_field)  # Aplica a ordenação com base no campo selecionado

    docentes = Docente.objects.all()
    cursos = Curso.objects.all()

    return render(request, 'cadastro_servidor.html', {'form': form, 'servidores': servidores, 'docentes': docentes, 'cursos': cursos})




def perfil_servidor(request, pk):
    servidor = Servidor.objects.get(pk=pk)
    documentos = Documento.objects.all
    return render(request, 'perfil_servidor.html', {'servidor' : servidor, 'documentos' : documentos})

def editar_servidor(request, pk):
    servidores = get_object_or_404(Servidor, pk=pk)
    if request.method == 'POST':
        form = ServidorForm(request.POST,instance=servidores)
        if form.is_valid():
            form.save()
            return redirect('cadastro_servidor')
    else:
        form = ServidorForm(instance=servidores)

    cursos = Curso.objects.all()  # Obtenha todos os cursos do banco de dados
    servidores = Servidor.objects.all()  # Obtenha todos os docentes do banco de dados

    return render(request, 'cadastro_servidor.html', {'form': form, 'cursos': cursos, 'servidores': servidores, })

from django.shortcuts import render, redirect
from .forms import DocumentoForm

def enviar_documento(request, pk):
    servidor = get_object_or_404(Servidor, pk=pk)
    if request.method == 'POST':
        form = DocumentoForm(request.POST, request.FILES)
        if form.is_valid():
            tipo = form.cleaned_data['tipo']

            # Verifica se já existe um documento com o mesmo tipo
            documento_existente = Documento.objects.filter(servidor=servidor, tipo=tipo).first()

            if documento_existente and not request.POST.get('substituir_documento'):
                return render(request, 'enviar_documento.html', {
                    'form': form,
                    'documento_existente': documento_existente,
                    'servidor': servidor
                })

            # Substitui o documento existente, se houver
            if documento_existente:
                documento_existente.arquivo.delete()
                documento_existente.delete()

            form.instance.servidor = servidor
            form.save()
            return redirect('documentos_enviados', pk=servidor.pk)
    else:
        form = DocumentoForm(initial={'servidor': servidor})

    return render(request, 'enviar_documento.html', {'form': form, 'servidor': servidor})



def documentos_enviados(request, pk):
    servidor = get_object_or_404(Servidor, pk=pk)
    documentos = Documento.objects.filter(servidor=servidor)
    return render(request, 'documentos_enviados.html', {'servidor': servidor, 'documentos': documentos})


def remover_documento(request, documento_id):
    documento = get_object_or_404(Documento, id=documento_id)
    servidor = documento.servidor
    documento.arquivo.delete()  # Remove o arquivo associado ao documento
    documento.delete()  # Remove o documento do banco de dados
    return redirect('documentos_enviados', pk=servidor.pk)

def substituir_documento(request, documento_id):
    documento_existente = get_object_or_404(Documento, id=documento_id)
    servidor = documento_existente.servidor

    if request.method == 'POST':
        form = DocumentoForm(request.POST, request.FILES, instance=documento_existente)
        if form.is_valid():
            form.save()
            return redirect('documentos_enviados', pk=servidor.pk)
    else:
        form = DocumentoForm(instance=documento_existente, use_required_attribute=False)

    return render(request, 'substituir_documento.html', {'servidor': servidor, 'form': form})



import qrcode
from io import BytesIO
from reportlab.lib.units import inch
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY
from PIL import Image as PilImage
import base64  # Importe base64

def emitir_declaracao(request):
    sucesso = False

    if request.method == 'POST':
        form = DeclaracaoEmitidaForm(request.POST)
        if form.is_valid():
            declaracao_emitida = DeclaracaoEmitida(
                codigo_autenticacao=gerar_codigo_unico(),
                docente=form.cleaned_data['docente'],
                curso=form.cleaned_data['curso'],
            )
            declaracao_emitida.save()

            # Gerar o QR code
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=10,
                border=4,
            )
            url_autenticacao = 'https://esprn.saude.rn.gov.br/ngtes/autenticar/'
            qr.add_data(url_autenticacao)
            qr.make(fit=True)

            qr_image = qr.make_image(fill_color="black", back_color="white")

            # Converter o QR code para base64
            buffered = BytesIO()
            qr_image.save(buffered, format="PNG")
            qr_image_base64 = base64.b64encode(buffered.getvalue()).decode("utf-8")

            buffer = BytesIO()
            doc = SimpleDocTemplate(
                buffer,
                pagesize=letter,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=18
            )

            conteudo = []
            styles = getSampleStyleSheet()

            logo_path = 'https://esprn.saude.rn.gov.br/extensao/imagens/logo.png'
            logo = Image(logo_path, width=2*inch, height=2*inch)
            conteudo.append(logo)

            estilo_ccabecalho = ParagraphStyle(
                'CabecalhoEstilo',
                parent=styles['Normal'], 
                alignment=TA_CENTER,
            )

            cabecalho = [
                Paragraph("<strong>Escola de Saúde Pública do Rio Grande do Norte</strong>", estilo_ccabecalho),
                Paragraph("<strong>Parecer CES/CEE-RN N°03/2021 31 de março de 2021</strong>", estilo_ccabecalho),
                Paragraph("<strong>Av. Alexandrino de Alencar, 1850 – Tirol – Natal/RN – CEP 59015-350 – Telefone (84) 3232 7634 / 7628</strong>", estilo_ccabecalho),
                Paragraph("<strong>www.esprn.rn.gov.br</strong>", estilo_ccabecalho),
                Spacer(1, 0.2 * inch),
                Paragraph("<strong>DECLARAÇÃO</strong>", style=ParagraphStyle(name='DeclaracaoEstilo', fontSize=16, alignment=TA_CENTER)),
                Spacer(1, 0.2 * inch),
            ]

            conteudo.extend(cabecalho)

            conteudo.append(Spacer(1, 0.2*inch))

            estilo_corpo = ParagraphStyle(
                'CorpoEstilo',
                parent=styles['Normal'], 
                alignment=TA_JUSTIFY,
            )

            corpo_declaracao = [
                Paragraph(f"Declaramos para os devidos fins que <strong>{declaracao_emitida.docente.nome}</strong>, inscrito(a) sob o CPF nº <strong>{declaracao_emitida.docente.cpf}</strong>, exerceu atividades como tutor(a) do curso <strong>{declaracao_emitida.curso.nome}</strong>, na modalidade semi-presencial, nesta Escola de Saúde Pública do Rio Grande do Norte - ESPRN, instituição integrante da Rede de Escolas Técnicas do SUS - RETSUS e da Rede Nacional de Escolas de Saúde Pública - RedEscola, perfazendo a carga horária total de {declaracao_emitida.curso.carga_horaria} horas.", estilo_corpo),
            ]

            conteudo.extend(corpo_declaracao)

            # Inserir o QR code no rodapé
            qr_image_base64_data = f"data:image/png;base64,{qr_image_base64}"
            qr_image_reportlab = Image(qr_image_base64_data, width=1.25*inch, height=1.25*inch)
            
            # Defina o estilo do rodapé
            rodape_style = ParagraphStyle(
                'RodapeEstilo',
                parent=styles['Normal'],
                alignment=TA_JUSTIFY,
            )
            
            # Crie o parágrafo com o código de autenticação
            rodape = [ 
                Paragraph(f"Código de Autenticação: <strong>{declaracao_emitida.codigo_autenticacao}</strong>", rodape_style),
                Paragraph(f"Leia o Qr Code acima, ou acesse https://esprn.saude.rn.gov.br/ngtes/autenticar/ , e insira o código de autenticação acima para verificar a autenticidade desse documento.</strong>", rodape_style),
            ]
            # Adicione espaço em branco
            conteudo.append(Spacer(1, 2.5 * inch))
            
            # Adicione o QR code e o código de autenticação no rodapé
            conteudo.append(Spacer(1, 0.2 * inch))
            conteudo.extend([qr_image_reportlab, rodape])

            doc.build(conteudo)

            buffer.seek(0)
            response = HttpResponse(buffer, content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename=declaracao.pdf'
            sucesso = True

            return response

    else:
        form = DeclaracaoEmitidaForm()

    declaracoes_emitidas = DeclaracaoEmitida.objects.all()

    return render(request, 'emitir_declaracao.html', {'form': form, 'sucesso': sucesso, 'declaracoes_emitidas': declaracoes_emitidas})



from .models import DeclaracaoEmitida

def autenticar_declaracao(request):
    if request.method == 'POST':
        codigo_autenticacao = request.POST.get('codigo_autenticacao')
        try:
            declaracao_emitida = DeclaracaoEmitida.objects.get(codigo_autenticacao=codigo_autenticacao)
            return render(request, 'declaracao_autenticada.html', {'declaracao_emitida': declaracao_emitida})
        except DeclaracaoEmitida.DoesNotExist:
            return render(request, 'codigo_inexistente.html')
    return render(request, 'autenticar_declaracao.html')

import random
import string

def gerar_codigo_unico(length=10):
    caracteres = string.ascii_letters + string.digits
    return ''.join(random.choice(caracteres) for _ in range(length))
